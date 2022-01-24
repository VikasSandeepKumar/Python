from openpyxl import Workbook
wb = Workbook()
ws = wb.active
# ws['C5'] = "GOOGLE"

# testdata = [['Name', 'City'],['Vikas','Hyderabad'],['Dheeraj','UK']]
# for data in testdata:
#     ws.append(data)

for i in range(1, 10):
    for j in range(1, 5):
        ws.cell(row=i, column=j).value=i+j
wb.save("demoexcel.xlsx")

