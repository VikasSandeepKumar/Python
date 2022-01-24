from openpyxl import Workbook, load_workbook

# wb = load_workbook(filename='demoexcel.xlsx')
# sh = wb.active
#
# print(sh['A3'].value)
# print(wb['DemoSheet']['A2'].value)

wb = load_workbook(filename='demoexcel.xlsx')
sh = wb['demoexcel']

row_ct = sh.max_row
col_ct = sh.max_column

# print(row_ct)
# print(col_ct)

for i in range(1, row_ct+1):
    for j in range(1, col_ct+1):
        print(sh.cell(row=i, column=j).value, end='')
    print('\n')




# print(sh['A2'].value)
#
# print(wb['DemoSheet']['A2'].value)
# print(sh['A3'].value)
# print(wb['DemoSheet']['A2'].value)
# print(sh.cell(row=2, column=2).value)